"""포스 장부 엑셀 → sales_daily / product_sales_daily 반영.

구글드라이브 데스크톱이 동기화해 둔 '장부관리' 폴더(로컬 경로)를 훑어서
새 파일·수정된 파일만 파싱한다. OAuth 불필요 — 사장님은 지금처럼 드라이브
월 폴더에 엑셀만 올리면 된다.

파일 유형은 **파일명이 아니라 내용**으로 판별한다(파일명은 매달 제각각이었음):
  · TOS(토스포스)  — "데이터 기준 시작일자" 머리말. 시트: 일별 결제합계
                     (매입사별 배민/쿠팡/요기요 컬럼 포함), 일별×상품 합계,
                     주문 상세, 상품 주문 상세.
  · IMU(아임유)    — "매출일시/영수증번호" 헤더의 건별 내역. 매장(키오스크)만.
                     2026.3까지 TOS와 병행 — 중복 없음(사장님 확인)이라 합산.
  · baemin/coupang 월 정산 엑셀 — TOS 도입 전(2025.10~12) 배달 일매출용.
  · 그 외(바로고·카드사·집계 리포트 등)는 조용히 무시.

같은 (파일명, 수정시각)은 pos_files 로그로 재파싱을 막고, 반영은 전부
(날짜, 채널/상품, 출처) upsert 라 몇 번을 다시 돌려도 안전하다.

시간대별(sales_hourly, schema_v11 · 2026-09-03): 매출 대시보드의 요일×시간대
히트맵용. TOS 는 '결제 상세내역' 시트(결제 한 건마다 결제시각·주문채널·매입사),
IMU 는 건별 내역의 매출일시에서 '시'를 뽑아 (날짜, 시, 채널, 출처)로 모은다.
파서는 (sales, products, hourly) 세 묶음을 돌려준다.
"""

import glob
import logging
import os
import re
from collections import defaultdict
from datetime import date, datetime

from database import mkt_store

logger = logging.getLogger(__name__)

DEFAULT_LEDGER_DIR = (
    r"C:\Users\명구\Google Drive\1. Project_현재진행하는일\1. Business"
    r"\베어글스_송도_타임스페이스\매출 세무관련\장부관리")


def ledger_dir() -> str:
    return os.getenv("MKT_LEDGER_DIR", DEFAULT_LEDGER_DIR)


# ---------------------------------------------------------------------------
# 공통 유틸
# ---------------------------------------------------------------------------

_DATE_RE = re.compile(r"^(\d{4})[-./](\d{1,2})[-./](\d{1,2})")


def _to_date(v):
    """셀 값 → date (datetime, 'YYYY-MM-DD', 'YYYYMMDD' 지원). 실패 시 None."""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v or "").strip()
    m = _DATE_RE.match(s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    if re.fullmatch(r"\d{8}", s):
        try:
            return date(int(s[:4]), int(s[4:6]), int(s[6:8]))
        except ValueError:
            return None
    return None


_TIME_RE = re.compile(r"(\d{1,2}):(\d{2})")


def _to_hour(v):
    """셀 값 → 0~23 시. datetime 이면 .hour, 'YYYY-MM-DD HH:MM:SS' 문자열이면
    시각 부분. 시각이 없으면 None (날짜만 있는 셀은 0시로 오해하면 안 된다)."""
    if isinstance(v, datetime):
        return v.hour
    if isinstance(v, date):
        return None
    s = str(v or "").strip()
    if not s:
        return None
    # 날짜 뒤에 붙은 시각만 본다 — '2026-08-31 22:42:01'
    tail = s[10:] if _DATE_RE.match(s) else s
    m = _TIME_RE.search(tail)
    if not m:
        return None
    h = int(m.group(1))
    return h if 0 <= h <= 23 else None


def _to_int(v):
    if v is None or v == "":
        return 0
    if isinstance(v, (int, float)):
        return int(round(v))
    s = re.sub(r"[,\s원]", "", str(v))
    try:
        return int(round(float(s)))
    except ValueError:
        return 0


def _cells(row):
    return ["" if c is None else str(c).strip() for c in row]


def _open(path):
    """read_only 워크북 열기 + dimension 보정.

    아임유(IMU) 엑셀은 dimension 레코드가 깨져 있어 그대로 iter_rows 하면
    행이 빈 셀 하나로 읽힌다 — reset_dimensions() 로 실제 셀을 세게 한다.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        try:
            ws.reset_dimensions()
        except AttributeError:
            pass
    return wb


# 매입사/컬럼명 → 채널 매핑 (TOS 일별 합계의 매입사별 컬럼)
_CHANNEL_MAP = {
    "배달의민족": "baemin", "PLUGIN_BAEMIN": "baemin",
    "쿠팡이츠": "coupang", "요기요": "yogiyo", "땡겨요": "ddangyo",
}


# ---------------------------------------------------------------------------
# 유형 판별
# ---------------------------------------------------------------------------

def detect_kind(wb) -> str:
    """시트 이름과 앞부분 셀에서 시그니처를 찾아 유형을 정한다.

    ⚠️ read_only 워크북은 이 탐지에서 일부 순회된 뒤라 다시 iter_rows 하면
    스트림이 깨질 수 있다 — 파싱은 반드시 새로 연 워크북으로 할 것
    (import_file 이 그렇게 한다).
    """
    names = "|".join(wb.sheetnames)
    if "결제 합계" in names or "데이터 기준" in names:
        return "tos"
    heads = []
    for ws in wb.worksheets[:3]:
        for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
            heads.extend(_cells(row))
    joined = "|".join(heads)
    if "매출 정산 기준" in joined:
        return "tos"
    if "매출일시" in joined and "영수증번호" in joined:
        return "imu"
    if "주문기준일자" in joined and "상품명" in joined:
        return "tos_products"          # TOS '상품 주문 상세'만 따로 내보낸 파일
    if "서비스거래번호" in joined:
        return "baemin_xls"
    if "쿠팡부담" in joined:
        return "coupang_xls"
    return "skip"


# ---------------------------------------------------------------------------
# TOS 파서
# ---------------------------------------------------------------------------

def _parse_tos_daily(ws):
    """일별 결제 합계 시트 → sales 행들.

    헤더 행: '결제 합계 기간', '결제금액', (부가세), '결제건수', ... '매입사별'
    다음 행: 매입사 이름들 (배달의민족/쿠팡이츠/... 컬럼 위치가 달마다 다름)

    ⚠️ 2026-08 장부부터는 헤더('기간','결제금액',…)와 매입사 이름이 **한 줄**에
    같이 있다(그 위 줄이 '결제수단별/매입사별' 묶음 제목). 옛 규칙대로 '다음
    행'에서 매입사를 찾으면 데이터 행을 읽어 채널 컬럼을 하나도 못 잡고,
    배달이 0 → 매장 = 총액으로 부풀려졌다(2026-09-03 발견, 8월 배민·쿠팡
    행이 통째로 비어 있었다). 헤더 줄에 매입사 이름이 있으면 그 줄을 쓴다.
    """
    rows = list(ws.iter_rows(values_only=True))
    head_i = None
    for i, row in enumerate(rows):
        cells = _cells(row)
        if "결제금액" in cells and ("기간" in cells or "결제건수" in cells):
            head_i = i
            break
    if head_i is None or head_i + 1 >= len(rows):
        return []
    head = _cells(rows[head_i])
    col_amount = head.index("결제금액") if "결제금액" in head else 1
    col_count = head.index("결제건수") if "결제건수" in head else None
    if any(name in _CHANNEL_MAP for name in head):
        sub, data_start = head, head_i + 1          # 한 줄 양식(2026-08~)
    else:
        sub, data_start = _cells(rows[head_i + 1]), head_i + 2
    chan_cols = {}
    for j, name in enumerate(sub):
        if name in _CHANNEL_MAP:
            chan_cols[j] = _CHANNEL_MAP[name]

    out = []
    for row in rows[data_start:]:
        d = _to_date(row[0] if row else None)
        if not d:
            continue
        total = _to_int(row[col_amount]) if col_amount < len(row) else 0
        count = (_to_int(row[col_count])
                 if col_count is not None and col_count < len(row) else None)
        delivery = 0
        chans = defaultdict(int)
        for j, ch in chan_cols.items():
            v = _to_int(row[j]) if j < len(row) else 0
            if v:
                chans[ch] += v
                delivery += v
        store = max(total - delivery, 0)
        out.append({"sale_date": str(d), "channel": "store",
                    "amount": store, "orders_count": count, "source": "tos"})
        for ch, v in chans.items():
            out.append({"sale_date": str(d), "channel": ch,
                        "amount": v, "orders_count": None, "source": "tos"})
    return out


def _parse_tos_product_summary(ws):
    """'상품 주문 합계' 시트 (기간, 상품명, 카테고리, 판매건수, ..., 실판매금액)"""
    rows = list(ws.iter_rows(values_only=True))
    head_i = None
    for i, row in enumerate(rows):
        cells = _cells(row)
        if "상품명" in cells and "판매건수" in cells:
            head_i = i
            break
    if head_i is None:
        return {}
    head = _cells(rows[head_i])
    ci_name = head.index("상품명")
    ci_cat = head.index("카테고리") if "카테고리" in head else None
    ci_qty = head.index("판매건수")
    ci_amt = next((j for j, h in enumerate(head) if h.startswith("실 판매 금액")
                   or h.startswith("실판매금액")), None)
    agg = {}
    for row in rows[head_i + 1:]:
        d = _to_date(row[0] if row else None)
        if not d or ci_name >= len(row):
            continue
        name = str(row[ci_name] or "").strip()
        if not name:
            continue
        key = (str(d), name)
        cur = agg.setdefault(key, {"qty": 0, "amount": 0,
                                   "category": (str(row[ci_cat]).strip()
                                                if ci_cat is not None and row[ci_cat]
                                                else None)})
        cur["qty"] += _to_int(row[ci_qty]) if ci_qty < len(row) else 0
        if ci_amt is not None and ci_amt < len(row):
            cur["amount"] += _to_int(row[ci_amt])
    return agg


def _parse_tos_product_detail(ws):
    """'상품 주문 상세내역' 시트 (주문기준일자, ..., 상품명, 수량, 실판매금액)"""
    rows = ws.iter_rows(values_only=True)
    head = None
    buffered = []
    for row in rows:
        cells = _cells(row)
        if "주문기준일자" in cells and "상품명" in cells:
            head = cells
            break
        buffered.append(row)
    if head is None:
        return {}
    ci_name = head.index("상품명")
    ci_cat = head.index("카테고리") if "카테고리" in head else None
    ci_qty = head.index("수량")
    ci_amt = next((j for j, h in enumerate(head)
                   if h.startswith("실판매금액") or h.startswith("실 판매")), None)
    agg = {}
    for row in rows:
        d = _to_date(row[0] if row else None)
        if not d or ci_name >= len(row):
            continue
        name = str(row[ci_name] or "").strip()
        if not name:
            continue
        key = (str(d), name)
        cur = agg.setdefault(key, {"qty": 0, "amount": 0,
                                   "category": (str(row[ci_cat]).strip()
                                                if ci_cat is not None
                                                and ci_cat < len(row) and row[ci_cat]
                                                else None)})
        cur["qty"] += _to_int(row[ci_qty]) if ci_qty < len(row) else 0
        if ci_amt is not None and ci_amt < len(row):
            cur["amount"] += _to_int(row[ci_amt])
    return agg


def _parse_tos_payment_detail(ws):
    """'결제 상세내역' 시트 → 시간대별 행들.

    컬럼: 결제기준일자, 결제시각, 주문채널(포스/키오스크/배달), 주문번호,
          결제건수, 결제금액, 부가세, 결제수단, 매입사, 결제상태, 결제취소시각
    · 채널: 주문채널이 '배달'이면 매입사(배달의민족/쿠팡이츠/요기요)로,
      나머지(포스·키오스크)는 매장(store).
    · 취소는 **음수 행**으로 따로 온다(원래 승인 행은 그대로 남는다). 그래서
      상태로 거르지 않고 금액을 부호째 합산한다 — 그래야 '결제 합계' 시트의
      일 총액과 원 단위로 맞는다(2026-08 실측: 승인 36,024,882 − 취소
      1,409,553 = 합계 34,615,329). 건수도 취소면 −1.
    """
    rows = ws.iter_rows(values_only=True)
    head = None
    for row in rows:
        cells = _cells(row)
        if "결제시각" in cells and "결제금액" in cells:
            head = cells
            break
    if head is None:
        return []
    ci_date = head.index("결제기준일자") if "결제기준일자" in head else 0
    ci_time = head.index("결제시각")
    ci_chan = head.index("주문채널") if "주문채널" in head else None
    ci_buyer = head.index("매입사") if "매입사" in head else None
    ci_amt = head.index("결제금액")
    agg = defaultdict(lambda: [0, 0])      # (date, hour, channel) -> [amount, count]
    for row in rows:
        if ci_time >= len(row):
            continue
        d = (_to_date(row[ci_date] if ci_date < len(row) else None)
             or _to_date(row[ci_time]))
        h = _to_hour(row[ci_time])
        if not d or h is None:
            continue
        amt = _to_int(row[ci_amt]) if ci_amt < len(row) else 0
        if not amt:
            continue
        chan_kind = (str(row[ci_chan] or "").strip()
                     if ci_chan is not None and ci_chan < len(row) else "")
        if "배달" in chan_kind:
            buyer = (str(row[ci_buyer] or "").strip()
                     if ci_buyer is not None and ci_buyer < len(row) else "")
            ch = _CHANNEL_MAP.get(buyer, "etc")
        else:
            ch = "store"
        agg[(str(d), h, ch)][0] += amt
        agg[(str(d), h, ch)][1] += 1 if amt > 0 else -1
    return [{"sale_date": d, "hour": h, "channel": ch,
             "amount": v[0], "orders_count": max(v[1], 0), "source": "tos"}
            for (d, h, ch), v in agg.items()]


def parse_tos(wb):
    """TOS 워크북 전체 → (sales_rows, product_rows, hourly_rows).

    시트 이름 기준(월 리포트: 데이터 기준/결제 합계/상품 주문 합계/
    결제 상세내역/상품 주문 상세내역), 이름이 다르면 헤더로 판별.
    """
    sales, prod_summary, prod_detail, hourly = [], {}, {}, []
    for ws in wb.worksheets:
        title = ws.title or ""
        if "결제 합계" in title:
            sales.extend(_parse_tos_daily(ws))
            continue
        if "상품 주문 합계" in title:
            prod_summary.update(_parse_tos_product_summary(ws))
            continue
        if "상품 주문 상세" in title:
            if not prod_summary:
                prod_detail.update(_parse_tos_product_detail(ws))
            continue
        if "결제 상세" in title:
            hourly.extend(_parse_tos_payment_detail(ws))
            continue
        if "데이터 기준" in title:
            continue
        heads = []
        for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
            heads.extend(_cells(row))
        joined = "|".join(heads)
        if "결제시각" in joined and "결제금액" in joined:
            hourly.extend(_parse_tos_payment_detail(ws))
        elif "결제금액" in joined and "결제건수" in joined:
            sales.extend(_parse_tos_daily(ws))
        elif "판매건수" in joined and "상품명" in joined:
            prod_summary.update(_parse_tos_product_summary(ws))
        elif "주문기준일자" in joined and "상품명" in joined:
            prod_detail.update(_parse_tos_product_detail(ws))
    # 합계 시트가 있으면 그걸 쓰고, 없으면(상세 전용 파일) 상세 집계를 쓴다
    agg = prod_summary or prod_detail
    products = [{"sale_date": d, "product": p, "category": v["category"],
                 "qty": v["qty"], "amount": v["amount"], "source": "tos"}
                for (d, p), v in agg.items()]
    return sales, products, hourly


# ---------------------------------------------------------------------------
# IMU(아임유) 파서 — 건별 내역 → 매장 일매출 + 상품별
# ---------------------------------------------------------------------------

def parse_imu(wb):
    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    head = None
    for row in rows:
        cells = _cells(row)
        if "매출일시" in cells and "메뉴 이름" in cells:
            head = cells
            break
    if head is None:
        return [], [], []
    ci_dt = head.index("매출일시")
    ci_name = head.index("메뉴 이름")
    ci_qty = head.index("수량")
    ci_item_amt = head.index("메뉴별 판매가") if "메뉴별 판매가" in head else None
    ci_sale_amt = head.index("매출금액") if "매출금액" in head else None

    daily = defaultdict(lambda: [0, 0])      # date -> [amount, receipts]
    prods = defaultdict(lambda: [0, 0])      # (date, name) -> [qty, amount]
    hours = defaultdict(lambda: [0, 0])      # (date, hour) -> [amount, receipts]
    for row in rows:
        if ci_dt >= len(row):
            continue
        d = _to_date(row[ci_dt])
        if not d:
            continue
        # 매출금액은 영수증 첫 행에만 있다 (환불행은 음수 — 그대로 합산해 상쇄)
        if ci_sale_amt is not None and ci_sale_amt < len(row):
            amt = _to_int(row[ci_sale_amt])
            if amt:
                daily[str(d)][0] += amt
                daily[str(d)][1] += 1 if amt > 0 else -1
                h = _to_hour(row[ci_dt])
                if h is not None:
                    hours[(str(d), h)][0] += amt
                    hours[(str(d), h)][1] += 1 if amt > 0 else -1
        name = str(row[ci_name] or "").strip() if ci_name < len(row) else ""
        if name:
            k = (str(d), name)
            prods[k][0] += _to_int(row[ci_qty]) if ci_qty < len(row) else 0
            if ci_item_amt is not None and ci_item_amt < len(row):
                prods[k][1] += _to_int(row[ci_item_amt])

    sales = [{"sale_date": d, "channel": "store", "amount": v[0],
              "orders_count": max(v[1], 0), "source": "imu"}
             for d, v in daily.items() if v[0] > 0]
    products = [{"sale_date": d, "product": p, "category": None,
                 "qty": q, "amount": a, "source": "imu"}
                for (d, p), (q, a) in prods.items() if q or a]
    hourly = [{"sale_date": d, "hour": h, "channel": "store",
               "amount": v[0], "orders_count": max(v[1], 0), "source": "imu"}
              for (d, h), v in hours.items() if v[0] > 0]
    return sales, products, hourly


# ---------------------------------------------------------------------------
# 배민/쿠팡 월 정산 파서 (TOS 도입 전 배달 일매출)
# ---------------------------------------------------------------------------

def parse_baemin_xls(wb):
    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    head = None
    for row in rows:
        cells = _cells(row)
        if "일자" in cells and "합계" in cells:
            head = cells
            break
    if head is None:
        return [], [], []
    ci_amt = head.index("합계")
    daily = defaultdict(lambda: [0, 0])
    for row in rows:
        d = _to_date(row[0] if row else None)
        if not d:
            continue
        daily[str(d)][0] += _to_int(row[ci_amt]) if ci_amt < len(row) else 0
        daily[str(d)][1] += 1
    return [{"sale_date": d, "channel": "baemin", "amount": v[0],
             "orders_count": v[1], "source": "baemin_xls"}
            for d, v in daily.items() if v[0]], [], []


def parse_coupang_xls(wb):
    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    head = None
    for row in rows:
        cells = _cells(row)
        if "일자" in cells and "주문금액" in cells:
            head = cells
            break
    if head is None:
        return [], [], []
    ci_amt = head.index("주문금액")
    ci_type = head.index("거래유형") if "거래유형" in head else None
    daily = defaultdict(lambda: [0, 0])
    for row in rows:
        d = _to_date(row[0] if row else None)
        if not d:
            continue
        amt = _to_int(row[ci_amt]) if ci_amt < len(row) else 0
        if ci_type is not None and ci_type < len(row) \
                and "취소" in str(row[ci_type] or ""):
            amt = -abs(amt)
        daily[str(d)][0] += amt
        daily[str(d)][1] += 1 if amt >= 0 else -1
    return [{"sale_date": d, "channel": "coupang", "amount": v[0],
             "orders_count": max(v[1], 0), "source": "coupang_xls"}
            for d, v in daily.items() if v[0]], [], []


_PARSERS = {"tos": parse_tos, "tos_products": parse_tos,
            "imu": parse_imu,
            "baemin_xls": parse_baemin_xls, "coupang_xls": parse_coupang_xls}


# ---------------------------------------------------------------------------
# 폴더 스캔 + 반영
# ---------------------------------------------------------------------------

def import_file(path) -> dict:
    """엑셀 1개 파싱·반영. 결과 요약 dict (kind, sales, products, from, to)."""
    name = os.path.basename(path)
    wb = _open(path)
    try:
        kind = detect_kind(wb)
    finally:
        wb.close()
    if kind == "skip" or kind not in _PARSERS:
        return {"kind": "skip", "sales": 0, "products": 0}
    # 탐지에 쓴 read_only 워크북은 스트림이 소모됐다 — 새로 열어 파싱한다
    wb = _open(path)
    try:
        sales, products, hourly = _PARSERS[kind](wb)
    finally:
        wb.close()
    dates = [r["sale_date"] for r in sales] or [r["sale_date"] for r in products]
    mkt_store.upsert_sales(sales)
    mkt_store.upsert_product_sales(products)
    try:
        mkt_store.upsert_sales_hourly(hourly)
    except Exception as e:  # noqa: BLE001 — schema_v11 미적용이어도 일매출은 살린다
        logger.warning("시간대별 반영 실패(%s): %s", name, e)
    logger.info("장부 반영: %s (%s) 매출 %d행, 상품 %d행, 시간대 %d행",
                name, kind, len(sales), len(products), len(hourly))
    return {"kind": kind if kind != "tos_products" else "tos",
            "sales": len(sales), "products": len(products),
            "hourly": len(hourly),
            "from": min(dates) if dates else None,
            "to": max(dates) if dates else None}


def scan_ledger(force=False) -> dict:
    """장부 폴더 전체를 훑어 새/변경 파일만 반영. 요약 dict 반환."""
    base = ledger_dir()
    if not os.path.isdir(base):
        return {"ok": False, "error": f"장부 폴더가 없습니다: {base}"}
    done, skipped, errors = [], 0, []
    paths = sorted(glob.glob(os.path.join(base, "**", "*.xls*"), recursive=True))
    for path in paths:
        name = os.path.basename(path)
        if name.startswith("~$"):
            continue
        try:
            st = os.stat(path)
            mtime = datetime.fromtimestamp(st.st_mtime).isoformat(timespec="seconds")
            if not force and mkt_store.pos_file_done(name, mtime):
                skipped += 1
                continue
            if path.lower().endswith(".xls"):   # 구형 xls(카드사 등)는 대상 아님
                mkt_store.log_pos_file(name, mtime, st.st_size, kind="skip")
                continue
            info = import_file(path)
            if info["kind"] != "skip" and not info["sales"] and not info["products"]:
                # 포스 파일로 판별됐는데 매출을 한 줄도 못 읽었다 — '반영 완료'로
                # 남기면 사장님이 [지금 반영]을 다시 눌러도 영영 안 읽힌다
                # (2026-08-30 감사). error 로 남겨 다음 스캔에 재시도되게 한다.
                mkt_store.log_pos_file(
                    name, mtime, st.st_size, kind=info["kind"],
                    status="error", note="파싱 결과 0행 — 장부 양식 확인 필요")
                errors.append(f"{name}: 매출을 한 줄도 못 읽음(양식 변경?)")
                continue
            mkt_store.log_pos_file(
                name, mtime, st.st_size, kind=info["kind"],
                date_from=info.get("from"), date_to=info.get("to"),
                note=(f"매출 {info['sales']}행, 상품 {info['products']}행, "
                      f"시간대 {info.get('hourly', 0)}행"))
            if info["kind"] != "skip":
                done.append(f"{name}({info['kind']})")
        except Exception as e:  # noqa: BLE001 — 파일 하나가 전체를 막지 않게
            logger.warning("장부 파일 실패 %s: %s", name, e)
            errors.append(f"{name}: {str(e)[:80]}")
            try:
                mkt_store.log_pos_file(name, mtime, st.st_size,
                                       status="error", note=str(e)[:200])
            except Exception:  # noqa: BLE001
                pass
    return {"ok": not errors, "imported": done, "skipped": skipped,
            "errors": errors}
